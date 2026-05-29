"""Generate Excel source documents for the demo."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Calibri", size=10)


def _style_header(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 4)


def generate_requirements():
    """Generate system_requirements.xlsx."""
    wb = Workbook()

    # Sheet 1: System Requirements
    ws = wb.active
    ws.title = "System Requirements"
    headers = ["Requirement ID", "Title", "Description", "Priority", "ASIL", "Subsystem",
               "Verification Method", "Status", "Trace To"]
    _style_header(ws, headers)

    reqs = [
        ["REQ-PM-001", "Power Mode State Machine", "PCM shall implement 6 power modes (OFF, ACC, RUN, CRANK, RUN+ENGINE, EMERGENCY) with defined transitions", "High", "ASIL-B", "PCM", "HIL Test", "Approved", "ES-PCM-2024-001"],
        ["REQ-PM-002", "Low Voltage Protection", "PCM shall transition to OFF when battery voltage < 9.0V for 30 continuous seconds", "High", "ASIL-B", "PCM", "HIL Test", "Approved", "ES-PCM-2024-001"],
        ["REQ-PM-003", "Overvoltage Protection", "PCM shall reduce alternator field when battery voltage > 16.0V for 10s", "High", "ASIL-B", "PCM", "HIL Test", "Approved", "ES-PCM-2024-001"],
        ["REQ-PM-004", "Calibration Integrity", "PCM shall verify calibration checksum on every power-on", "Critical", "ASIL-C", "PCM", "Unit Test", "Approved", "ES-PCM-2024-001"],
        ["REQ-TR-001", "Gear Range Selection", "TCM shall support P, R, N, D ranges with defined interlocks", "High", "ASIL-B", "TCM", "HIL Test", "Approved", "ES-TRANS-2024-002"],
        ["REQ-TR-002", "Shift Timing", "Total shift time shall not exceed 1.2 seconds", "High", "ASIL-B", "TCM", "HIL Test", "Approved", "ES-TRANS-2024-002"],
        ["REQ-TR-003", "Shift Quality", "Acceleration discontinuity during shift < 0.3g", "Medium", "ASIL-A", "TCM", "HIL Test", "Approved", "ES-TRANS-2024-002"],
        ["REQ-TR-009", "Shift Rejection", "TCM shall reject shift if speed > 5 mph for reverse engagement", "High", "ASIL-B", "TCM", "HIL Test", "Approved", "ES-TRANS-2024-002"],
        ["REQ-TR-010", "Range Sensor Validation", "TCM shall detect range sensor conflict within 500ms", "High", "ASIL-B", "TCM", "HIL Test", "Approved", "ES-TRANS-2024-002"],
        ["REQ-TR-011", "Gear Ratio Monitor", "TCM shall detect gear ratio deviation > 15%", "High", "ASIL-B", "TCM", "HIL Test", "Approved", "ES-TRANS-2024-002"],
        ["REQ-NET-001", "ECM Communication", "PCM shall detect lost ECM communication within 1 second", "High", "ASIL-B", "Network", "HIL Test", "Approved", "ES-CAN-2024-003"],
        ["REQ-NET-002", "Bus Off Recovery", "Gateway shall attempt CAN bus recovery on bus-off condition", "Medium", "ASIL-A", "Network", "HIL Test", "Approved", "ES-CAN-2024-003"],
        ["REQ-EM-001", "OBD Monitor Status", "PCM shall track OBD-II monitor completion status", "Medium", "QM", "Emissions", "Bench Test", "Approved", "ES-DTC-2024-004"],
    ]

    for row_idx, req in enumerate(reqs, 2):
        for col_idx, val in enumerate(req, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT

    # Sheet 2: Verification Matrix
    ws2 = wb.create_sheet("Verification Matrix")
    headers2 = ["Requirement ID", "Test Method", "Test ID", "Expected Result", "Pass Criteria"]
    _style_header(ws2, headers2)

    verif = [
        ["REQ-PM-001", "HIL", "TC-PM-001", "All 6 modes reachable", "All transitions execute within timing spec"],
        ["REQ-PM-002", "HIL", "TC-PM-002", "Shutdown at < 9.0V", "Shutdown within 30s ± 1s"],
        ["REQ-TR-001", "HIL", "TC-TR-001", "All ranges selectable", "P/R/N/D engage correctly"],
        ["REQ-TR-002", "HIL", "TC-TR-002", "Shift time < 1.2s", "Measured shift time under limit"],
        ["REQ-NET-001", "HIL", "TC-NET-001", "Detection within 1s", "DTC U0100 set within 1s ± 100ms"],
    ]
    for row_idx, row in enumerate(verif, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT

    wb.save("source_documents/excel/system_requirements.xlsx")
    print("  Generated: system_requirements.xlsx")


def generate_parameters():
    """Generate test_parameters.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Calibration Parameters"

    headers = ["Parameter ID", "Name", "Subsystem", "Min Value", "Max Value",
               "Default", "Unit", "Description"]
    _style_header(ws, headers)

    params = [
        ["CAL-001", "BATT_V_LOW_THRESH", "PCM", 8.5, 9.5, 9.0, "V", "Low voltage shutdown threshold"],
        ["CAL-002", "BATT_V_HIGH_THRESH", "PCM", 15.5, 16.5, 16.0, "V", "Overvoltage detection threshold"],
        ["CAL-003", "CRANK_TIMEOUT", "PCM", 8.0, 15.0, 10.0, "s", "Maximum cranking duration"],
        ["CAL-004", "ENGINE_STALL_RPM", "PCM", 150, 300, 200, "rpm", "Stall detection RPM threshold"],
        ["CAL-005", "ENGINE_STALL_TIME", "PCM", 1.0, 3.0, 2.0, "s", "Stall detection duration"],
        ["CAL-006", "ENGINE_START_RPM", "PCM", 350, 500, 400, "rpm", "Successful start RPM threshold"],
        ["CAL-007", "SHIFT_SOLENOID_RESP", "TCM", 30, 80, 50, "ms", "Solenoid response time limit"],
        ["CAL-008", "GEAR_ENGAGE_TIME", "TCM", 200, 400, 300, "ms", "Gear engagement time limit"],
        ["CAL-009", "TC_LOCKUP_TIME", "TCM", 300, 700, 500, "ms", "Torque converter lockup time"],
        ["CAL-010", "TOTAL_SHIFT_TIME", "TCM", 800, 1500, 1200, "ms", "Total shift duration limit"],
        ["CAL-011", "SHIFT_QUALITY_G", "TCM", 0.2, 0.5, 0.3, "g", "Max acceleration discontinuity"],
        ["CAL-012", "REVERSE_SPEED_LIM", "TCM", 3, 8, 5, "mph", "Max speed for reverse engagement"],
        ["CAL-013", "PARK_SPEED_LIM", "TCM", 2, 5, 3, "mph", "Max speed for park engagement"],
        ["CAL-014", "CAN_TIMEOUT", "Network", 500, 2000, 1000, "ms", "CAN message timeout threshold"],
        ["CAL-015", "OIL_PRESS_MIN", "PCM", 3, 8, 5, "psi", "Minimum oil pressure threshold"],
        ["CAL-016", "COOLANT_TEMP_MAX", "PCM", 110, 130, 120, "C", "Maximum coolant temperature"],
    ]

    for row_idx, param in enumerate(params, 2):
        for col_idx, val in enumerate(param, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT

    wb.save("source_documents/excel/test_parameters.xlsx")
    print("  Generated: test_parameters.xlsx")


if __name__ == "__main__":
    print("Generating Excel source documents...")
    generate_requirements()
    generate_parameters()
    print("Done.")
