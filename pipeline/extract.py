"""Stage 2: Extract — identify text, tables, diagram regions, and structured data."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from pipeline.models import DocumentJob, PipelineStage


def extract_content(job: DocumentJob, file_path: Path) -> dict[str, Any]:
    """Extract structured content from a document file.

    For pre-extracted JSON files, loads them directly.
    For markdown files, identifies sections and embedded references.
    For Excel files, reads sheet structure.
    """
    job.advance(PipelineStage.EXTRACTING, f"Extracting content from {job.filename}")

    extracted: dict[str, Any] = {"raw_sections": [], "tables": 0, "diagrams": [], "entities": {}}

    if job.file_type == "extracted_json":
        extracted = _extract_from_json(file_path)
    elif job.file_type == "markdown_spec":
        extracted = _extract_from_markdown(file_path)
    elif job.file_type == "excel_workbook":
        extracted = _extract_from_excel(file_path)
    elif job.file_type == "pdf_document":
        extracted = _extract_from_pdf(file_path)
    else:
        extracted["raw_sections"] = [{"type": "unknown", "content": "Unsupported file type"}]

    entity_count = sum(
        len(v) if isinstance(v, list) else 1
        for v in extracted.get("entities", {}).values()
    )
    job.advance(
        PipelineStage.EXTRACTING,
        f"Extracted {entity_count} entities, {extracted.get('tables', 0)} tables, "
        f"{len(extracted.get('diagrams', []))} diagrams",
    )
    return extracted


def _extract_from_json(file_path: Path) -> dict[str, Any]:
    """Extract from pre-extracted JSON files."""
    with open(file_path) as f:
        data = json.load(f)

    entities: dict[str, list[Any]] = {}
    diagrams: list[str] = data.get("diagrams", [])
    tables = data.get("tables", 0)

    # Collect all entity arrays from the JSON
    for key in ["states", "transitions", "signals", "dtcs", "input_signals",
                 "output_signals", "gear_ranges", "safety_interlocks",
                 "messages", "upshift_schedule"]:
        if key in data:
            val = data[key]
            if isinstance(val, list):
                entities[key] = val
            elif isinstance(val, dict) and "shifts" in val:
                entities[key] = val["shifts"]

    return {
        "raw_sections": [],
        "tables": tables,
        "diagrams": diagrams,
        "entities": entities,
        "metadata": {
            k: data[k]
            for k in ["document_id", "title", "revision", "subsystem", "asil",
                       "effective_date"]
            if k in data
        },
    }


def _extract_from_markdown(file_path: Path) -> dict[str, Any]:
    """Extract structure from markdown spec files."""
    content = file_path.read_text()
    sections: list[dict[str, str]] = []
    diagrams: list[str] = []
    table_count = 0

    current_section = ""
    for line in content.split("\n"):
        if line.startswith("#"):
            current_section = line.lstrip("#").strip()
            sections.append({"heading": current_section, "type": "section"})
        elif "![" in line:
            # Image reference
            start = line.find("(") + 1
            end = line.find(")")
            if start > 0 and end > start:
                diagrams.append(line[start:end])
        elif line.startswith("|") and "---" not in line:
            table_count += 1

    return {
        "raw_sections": sections,
        "tables": max(1, table_count // 3),
        "diagrams": diagrams,
        "entities": {},
    }


def _extract_from_excel(file_path: Path) -> dict[str, Any]:
    """Extract from Excel workbooks."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        entities: dict[str, list[dict[str, Any]]] = {}
        table_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h or "").strip() for h in rows[0]]
            records: list[dict[str, Any]] = []
            for row in rows[1:]:
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        record[headers[i]] = val
                if any(v is not None for v in record.values()):
                    records.append(record)
            if records:
                entities[sheet_name.lower().replace(" ", "_")] = records
                table_count += 1

        wb.close()
        return {
            "raw_sections": [{"type": "sheet", "heading": s} for s in wb.sheetnames],
            "tables": table_count,
            "diagrams": [],
            "entities": entities,
        }
    except Exception:
        return {"raw_sections": [], "tables": 0, "diagrams": [], "entities": {}}


def _extract_from_pdf(file_path: Path) -> dict[str, Any]:
    """Extract tables, metadata, and diagram regions from a PDF spec.

    Real unstructured-document parsing: ``pdfplumber`` pulls ruled tables and
    page text out of the binary PDF, then table headers are classified into
    data-lake entity groups (DTCs, CAN signals, calibration parameters).
    """
    try:
        import pdfplumber
    except ImportError:
        return {"raw_sections": [], "tables": 0, "diagrams": [], "entities": {}}

    entities: dict[str, list[dict[str, Any]]] = {}
    diagrams: list[str] = []
    table_count = 0
    text_parts: list[str] = []
    headings: list[dict[str, str]] = []

    with pdfplumber.open(str(file_path)) as pdf:
        for page_index, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            text_parts.append(text)

            # Vector diagrams (boxes/arrows) show up as curves; ruled tables do not.
            if page.curves:
                diagrams.append(f"page-{page_index + 1}-diagram")

            for table in page.extract_tables():
                if not table or len(table) < 2:
                    continue
                header = [_clean_cell(h) for h in table[0]]
                if not any(header):
                    continue
                rows = [
                    r for r in table[1:]
                    if any(_clean_cell(c) for c in r)
                ]
                if not rows:
                    continue
                table_count += 1
                category_key, mapper = _classify_table(header)
                if category_key is None:
                    continue
                for raw_row in rows:
                    record = {
                        header[i]: _clean_cell(raw_row[i])
                        for i in range(len(header))
                        if i < len(raw_row) and header[i]
                    }
                    entities.setdefault(category_key, []).append(mapper(record))

    for line in "\n".join(text_parts).split("\n"):
        if line.strip().split(".")[0].isdigit() and len(line.strip()) < 80:
            headings.append({"heading": line.strip(), "type": "section"})

    return {
        "raw_sections": headings,
        "tables": table_count,
        "diagrams": diagrams,
        "entities": entities,
        "metadata": _parse_pdf_metadata("\n".join(text_parts), file_path),
    }


def _clean_cell(value: Any) -> str:
    """Normalize a table cell to a single-line trimmed string."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _pick(record: dict[str, str], *keywords: str) -> str:
    """Return the first cell whose header contains any of the keywords."""
    for key, val in record.items():
        low = key.lower()
        if any(kw in low for kw in keywords):
            return val
    return ""


def _num(value: str) -> Any:
    """Convert a numeric string to int/float, otherwise return it unchanged."""
    if not value:
        return value
    try:
        f = float(value)
        return int(f) if f.is_integer() else f
    except ValueError:
        return value


def _classify_table(
    header: list[str],
) -> tuple[str | None, Any]:
    """Map a table header to a data-lake category key and a row mapper."""
    joined = " ".join(h.lower() for h in header)

    if "dtc" in joined or ("code" in joined and "fault" in joined):
        return "dtcs", _map_dtc
    if "signal" in joined and ("message" in joined or "bit" in joined):
        return "signals", _map_signal
    if "parameter" in joined or ("min" in joined and "max" in joined):
        return "parameters", _map_parameter
    return None, None


def _map_dtc(record: dict[str, str]) -> dict[str, Any]:
    return {
        "code": _pick(record, "code"),
        "description": _pick(record, "description", "desc"),
        "fault_action": _pick(record, "action", "fault"),
        "mil": _pick(record, "mil"),
        "debounce_ms": _num(_pick(record, "debounce")),
    }


def _map_signal(record: dict[str, str]) -> dict[str, Any]:
    return {
        "name": _pick(record, "signal name", "name"),
        "message_id": _pick(record, "message"),
        "start_bit": _num(_pick(record, "start")),
        "length": _num(_pick(record, "length")),
        "scale": _num(_pick(record, "scale")),
        "unit": _pick(record, "unit"),
        "cycle_time_ms": _num(_pick(record, "cycle")),
    }


def _map_parameter(record: dict[str, str]) -> dict[str, Any]:
    # Preserve original header keys so downstream structuring finds "Parameter ID".
    return dict(record)


def _parse_pdf_metadata(text: str, file_path: Path) -> dict[str, Any]:
    """Pull document identity fields out of the PDF cover text."""
    import re

    metadata: dict[str, Any] = {}
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    if lines:
        metadata["title"] = lines[0]

    patterns = {
        "document_id": r"Document ID:\s*([A-Za-z0-9\-]+)",
        "revision": r"Revision:\s*([A-Za-z0-9.]+)",
        "subsystem": r"Subsystem:\s*([^|\n]+)",
        "asil": r"(ASIL\s*[A-D])",
        "effective_date": r"Effective Date:\s*([0-9\-/]+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            metadata[key] = match.group(1).strip()

    metadata.setdefault("document_id", file_path.stem)
    return metadata
